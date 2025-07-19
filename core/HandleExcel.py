import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class CHandleExcel:
    def __init__(self, file_path):
        """
        初始化函数，设置文件路径。

        :param file_path: Excel 文件的路径
        """
        self.file_path = file_path

    def find_empty_start_column(self):
        """
        查找Excel中第一列为空的起始列，返回列的索引。

        该函数会读取Excel文件，检查每一列的数据，如果某一列的数据全部为空或为空白字符串，
        则返回该列的索引。否则，返回Excel文件中所有列之后的下一列索引。

        :return: (bool, int) 成功时返回 (True, 列号)，失败时返回 (False, 错误信息)
        """
        try:
            df = pd.read_excel(self.file_path)  # 使用 pandas 读取 Excel 文件
            num_columns = len(df.columns)  # 获取列数
            for i in range(num_columns):
                column_data = df.iloc[:, i]
                # 如果该列为空或全部为空白字符，返回当前列的索引（+2 是因为 Excel 的列索引从 1 开始）
                if column_data.isnull().all() or all(str(value).strip() == '' for value in column_data):
                    return True, i + 2
            # 如果没有找到空列，返回最后一列的下一列
            return True, num_columns + 2
        except Exception as e:
            print(f'读取Excel文件出错：{str(e)}')
            return False, f'读取Excel文件出错：{str(e)}'

    def read_urls_from_excel(self, column_name=False):
        """
        从Excel文件中读取URL列表。

        该函数会尝试读取名为 "url" 或 "link" 的列，并返回这些列中的数据。用户也可以指定
        自定义列名来查找对应的列。

        :param column_name: 自定义列名，默认为 False
        :return: (bool, dict) 成功时返回 (True, urls_dict)，失败时返回 (False, 错误信息)
        """
        try:
            df = pd.read_excel(self.file_path)
            url_column = None
            if column_name:
                for col in df.columns:
                    if col.lower() == column_name:
                        url_column = col
                        break
            else:
                for col in df.columns:
                    if col.lower() == "url":
                        url_column = col
                        break
                    elif col.lower() == 'link':
                        url_column = col
                        break
            if url_column is None:
                print("未找到包含关键字的列的列")
                return False, "未找到包含关键字的列的列"

            # 将对应列的URL数据存储为字典，键为行号，值为URL
            urls_dict = {}
            for index, value in df[url_column].items():
                urls_dict[index + 2] = value  # 行号从0开始，所以要加上2

            return True, urls_dict

        except Exception as e:
            print(f'读取Excel文件出错：{str(e)}')
            return False, f'读取Excel文件出错：{str(e)}'

    def write_excel_header(self, header_list, start_col):
        """
        向Excel表格的指定位置写入表头。

        该函数接受一个表头列表，并将其写入Excel表格的第一行，从指定的列开始。
        同时将表头设置为粗体。

        :param header_list: 包含表头数据的列表
        :param start_col: 表头写入的起始列号
        :return: bool 成功时返回 True，失败时返回 False
        """
        try:
            wb = load_workbook(self.file_path)  # 加载Excel文件
            sheet = wb.active  # 获取当前活动工作表
            # 写入表头
            for idx, header in enumerate(header_list):
                col_letter = get_column_letter(start_col + idx)  # 获取列字母
                sheet[f"{col_letter}1"] = header  # 将表头写入第一行

            # 设置表头为粗体
            for col in range(1, start_col + len(header_list)):
                col_letter = get_column_letter(col)
                cell = sheet[f"{col_letter}1"]
                cell.font = Font(bold=True)

            wb.save(self.file_path)  # 保存修改后的文件
            return True
        except Exception as e:
            print(f"写入Excel文件出错：{e}")
            return False

    def write_excel_data(self, data_list, start_col):
        """
        向Excel表格中写入数据。

        该函数接受一个字典格式的数据，其中键为行号，值为该行的数据列表。
        数据将从指定的起始列开始写入。

        :param data_list: 包含数据的字典，键为行号，值为数据列表
        :param start_col: 数据写入的起始列号
        :return: bool 成功时返回 True，失败时返回 False
        """
        try:
            wb = load_workbook(self.file_path)  # 加载Excel文件
            sheet = wb.active  # 获取当前活动工作表

            # 遍历数据并写入Excel
            for row_letter in data_list:
                data_list_row = data_list[row_letter]
                for idx, data_li in enumerate(data_list_row):
                    col_letter = get_column_letter(start_col + idx)  # 获取列字母
                    sheet[f"{col_letter}{row_letter}"] = data_li  # 将数据写入单元格

            wb.save(self.file_path)  # 保存修改后的文件
            return True
        except Exception as e:
            print(f"写入Excel文件出错：{e}")
            return False
